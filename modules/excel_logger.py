"""openpyxl Excel project logger."""
import os
from datetime import datetime
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent / "video_log.xlsx"

HEADERS = [
    "Date Created", "Topic", "Best Title", "Duration", "Quality",
    "Style", "Voice", "Content Type", "Video File Path", "Thumbnail Path",
    "Tags (CSV)", "Description (200 chars)", "Upload Status",
    "YouTube URL", "Views (7-day)", "Revenue ($)",
]


def _init_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Video Log"

    # Header row
    ws.append(HEADERS)
    for col, cell in enumerate(ws[1], 1):
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor="1a1a1a")
        cell.alignment = Alignment(horizontal="center")

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(str(EXCEL_PATH))
    return wb


def log_project(project: dict) -> None:
    """Append one row to video_log.xlsx. Auto-creates the file if needed."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    if not EXCEL_PATH.exists():
        _init_workbook()

    wb = load_workbook(str(EXCEL_PATH))
    ws = wb.active

    tags_csv = ", ".join(project.get("tags", [])[:20])
    desc = (project.get("description", ""))[:200]

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        project.get("topic", ""),
        project.get("best_title", ""),
        project.get("duration", ""),
        project.get("quality", ""),
        project.get("style", ""),
        project.get("voice", ""),
        project.get("content_type", ""),
        project.get("video_path", ""),
        project.get("thumbnail_path", ""),
        tags_csv,
        desc,
        "",  # Upload Status — manual
        "",  # YouTube URL — manual
        "",  # Views
        "",  # Revenue
    ]

    # Insert after header (most recent first)
    ws.insert_rows(2)
    for col, value in enumerate(row, 1):
        ws.cell(row=2, column=col, value=value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(50, max(12, max_len + 2))

    wb.save(str(EXCEL_PATH))

    # Open the file automatically
    try:
        os.startfile(str(EXCEL_PATH))
    except Exception:
        pass
